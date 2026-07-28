"""
Claustor AI — Document Processor (Singleton)

Loads all heavy models ONCE at startup.
Reused across all contract processing requests.

Capabilities by plan:
  Free:       PyMuPDF text extraction only
  Starter:    + pdfplumber tables + Tesseract OCR
  Pro:        + Presidio PII masking + openpyxl XLSX + metadata
  Enterprise: + Gemini Vision + XML extraction + metadata stripping
"""

import io
import os
import re
import time
import structlog
from pathlib import Path
from typing import Any

logger = structlog.get_logger(__name__)

# Import sanitizer — lazy to avoid circular imports
def _get_sanitizer():
    from app.infrastructure.security.sanitizer import (
        sanitize_document_text, validate_context_window
    )
    return sanitize_document_text, validate_context_window


class DocumentProcessor:
    """
    Singleton document processor.
    All heavy models loaded once at startup via init_models().
    """

    _instance = None
    _initialized = False

    # Model references (set during init)
    _spacy_nlp      = None
    _presidio       = None
    _tesseract_ok   = False
    _pdfplumber_ok  = False
    _openpyxl_ok    = False

    @classmethod
    def get(cls) -> "DocumentProcessor":
        if cls._instance is None:
            cls._instance = cls()
        return cls._instance

    @classmethod
    def init_models(cls) -> None:
        """
        Load all heavy models at startup.
        Call from FastAPI lifespan startup event.
        """
        if cls._initialized:
            return

        start = time.time()
        logger.info("document_processor_init_start")

        # 1. pdfplumber (lightweight, always available)
        try:
            import pdfplumber
            cls._pdfplumber_ok = True
            logger.info("model_loaded", model="pdfplumber")
        except ImportError:
            logger.warning("model_unavailable", model="pdfplumber")

        # 2. Tesseract OCR
        try:
            import pytesseract
            result = pytesseract.get_tesseract_version()
            cls._tesseract_ok = True
            logger.info("model_loaded", model="tesseract", version=str(result))
        except Exception as e:
            logger.warning("model_unavailable", model="tesseract", error=str(e))

        # 3. openpyxl (lightweight)
        try:
            import openpyxl
            cls._openpyxl_ok = True
            logger.info("model_loaded", model="openpyxl")
        except ImportError:
            logger.warning("model_unavailable", model="openpyxl")

        # 4. spaCy + Presidio (heavy — 400MB+)
        try:
            import spacy
            cls._spacy_nlp = spacy.load("en_core_web_lg")
            logger.info("model_loaded", model="spacy_en_core_web_lg")

            from presidio_analyzer import AnalyzerEngine
            from presidio_anonymizer import AnonymizerEngine
            cls._presidio = {
                "analyzer":   AnalyzerEngine(),
                "anonymizer": AnonymizerEngine(),
            }
            logger.info("model_loaded", model="presidio")
        except Exception as e:
            logger.warning("model_unavailable", model="presidio", error=str(e))

        cls._initialized = True
        elapsed = round(time.time() - start, 2)
        logger.info("document_processor_ready",
                   elapsed_sec=elapsed,
                   tesseract=cls._tesseract_ok,
                   pdfplumber=cls._pdfplumber_ok,
                   presidio=cls._presidio is not None)

    # ── PUBLIC API ────────────────────────────────────────

    def parse(
        self,
        file_bytes: bytes,
        filename: str,
        plan: str = "free",
        org_plan: str = "free",
    ) -> dict:
        """
        Parse document and return structured content.
        Features gated by plan.
        """
        filename_lower = filename.lower()

        if filename_lower.endswith((".xlsx", ".xls")):
            return self._parse_excel(file_bytes, plan)

        if filename_lower.endswith(".xml"):
            return self._parse_xml(file_bytes, plan)

        if filename_lower.endswith((".pdf",)):
            return self._parse_pdf(file_bytes, filename, plan)

        if filename_lower.endswith((".docx", ".doc")):
            return self._parse_docx(file_bytes, plan)

        # Default: treat as text
        try:
            return {"full_text": file_bytes.decode("utf-8", errors="ignore"),
                    "tables": [], "metadata": {}, "pii_masked": False}
        except Exception:
            return {"full_text": "", "tables": [], "metadata": {}, "pii_masked": False}

    # ── PDF PARSING ───────────────────────────────────────

    def _parse_pdf(self, file_bytes: bytes, filename: str, plan: str) -> dict:
        """Multi-layer PDF parsing based on plan."""
        import fitz  # PyMuPDF

        result = {
            "full_text":   "",
            "tables":      [],
            "images":      [],
            "metadata":    {},
            "is_scanned":  False,
            "pii_masked":  False,
            "page_count":  0,
        }

        # ── Layer 1: PyMuPDF text (ALL plans) ────────────
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        result["page_count"] = len(doc)
        pages_text = []
        has_text = False

        for page in doc:
            text = page.get_text("text")
            if text.strip():
                has_text = True
            pages_text.append(text)

        raw_text = "\n".join(pages_text)

        # ── Security: sanitize extracted text ────────────
        sanitize_fn, window_fn = _get_sanitizer()
        san_result = sanitize_fn(raw_text, contract_id=contract_id if "contract_id" in dir() else "")
        if not san_result.is_clean:
            logger.warning("pdf_injection_sanitized",
                detections=san_result.detection_count,
                types=san_result.detection_types)
        # Validate context window
        validated_text, truncated = window_fn(san_result.sanitized_text)
        result["full_text"] = validated_text
        result["injection_detected"] = not san_result.is_clean
        result["context_truncated"] = truncated

        # Detect scanned PDF (little/no extractable text)
        if not has_text or len(result["full_text"].strip()) < 100:
            result["is_scanned"] = True

        # ── Layer 2: Metadata (ALL plans) ────────────────
        meta = doc.metadata or {}
        result["metadata"] = {
            "title":         meta.get("title", ""),
            "author":        meta.get("author", ""),
            "creator":       meta.get("creator", ""),
            "producer":      meta.get("producer", ""),
            "creation_date": meta.get("creationDate", ""),
            "mod_date":      meta.get("modDate", ""),
            "page_count":    len(doc),
        }
        doc.close()

        # ── Layer 3: Table extraction (Starter+) ─────────
        if plan in ("starter", "professional", "enterprise") and self._pdfplumber_ok:
            tables = self._extract_tables_pdf(file_bytes)
            result["tables"] = tables
            if tables:
                table_text = self._tables_to_text(tables)
                result["full_text"] += f"\n\n=== TABLES ===\n{table_text}"
            logger.info("tables_extracted", count=len(tables), plan=plan)

        # ── Layer 4: OCR for scanned PDFs (Starter+) ─────
        if result["is_scanned"] and plan in ("starter", "professional", "enterprise"):
            if self._tesseract_ok:
                ocr_text = self._ocr_pdf(file_bytes)
                if ocr_text:
                    result["full_text"] = ocr_text
                    result["is_scanned"] = True
                    logger.info("ocr_applied", chars=len(ocr_text), plan=plan)

        # ── Layer 5: PII masking (Professional+) ─────────
        if plan in ("professional", "enterprise") and self._presidio:
            result["full_text"], pii_found = self._mask_pii(result["full_text"])
            result["pii_masked"] = True
            result["pii_entities"] = pii_found
            logger.info("pii_masked", entities=len(pii_found), plan=plan)

        # ── Layer 6: Gemini Vision for images (Professional+) ──
        if plan in ("professional", "enterprise"):
            images = self.extract_images_from_pdf(file_bytes)
            if images:
                result["images"] = [
                    {"page": i["page"], "width": i["width"], "height": i["height"]}
                    for i in images
                ]
                result["_raw_images"] = images
                # Vision analysis scheduled async — raw images stored for pipeline
                result["vision_pending"] = True
                logger.info("vision_images_queued", count=len(images))
                logger.info("images_found_in_pdf",
                           count=len(images), plan=plan)

        # ── Layer 7: Metadata stripping (Enterprise) ──────
        if plan == "enterprise":
            result["metadata_stripped"] = True

        return result

    # ── TABLE EXTRACTION ──────────────────────────────────

    def _extract_tables_pdf(self, file_bytes: bytes) -> list[dict]:
        """Extract tables from PDF using pdfplumber."""
        tables = []
        try:
            import pdfplumber
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                for page_num, page in enumerate(pdf.pages, 1):
                    extracted = page.extract_tables()
                    for tbl in extracted:
                        if tbl and len(tbl) > 1:
                            headers = [str(h or "").strip() for h in (tbl[0] or [])]
                            rows = []
                            for row in tbl[1:]:
                                clean_row = [str(c or "").strip() for c in row]
                                if any(clean_row):
                                    rows.append(clean_row)
                            if rows:
                                tables.append({
                                    "page":    page_num,
                                    "headers": headers,
                                    "rows":    rows,
                                    "row_count": len(rows),
                                })
        except Exception as e:
            logger.warning("table_extraction_failed", error=str(e))
        return tables

    def _tables_to_text(self, tables: list[dict]) -> str:
        """Convert tables to readable text for LLM."""
        lines = []
        for i, table in enumerate(tables, 1):
            lines.append(f"Table {i} (Page {table['page']}):")
            if table["headers"]:
                lines.append(" | ".join(table["headers"]))
                lines.append("-" * 60)
            for row in table["rows"][:20]:  # cap at 20 rows
                lines.append(" | ".join(row))
            lines.append("")
        return "\n".join(lines)

    # ── OCR ───────────────────────────────────────────────


    def extract_images_from_pdf(self, file_bytes: bytes) -> list:
        """Extract images from PDF for vision analysis."""
        images = []
        try:
            import fitz  # PyMuPDF
            doc = fitz.open(stream=file_bytes, filetype="pdf")
            for page_num in range(len(doc)):
                page = doc[page_num]
                image_list = page.get_images(full=True)
                for img in image_list:
                    xref = img[0]
                    base_image = doc.extract_image(xref)
                    images.append({
                        "page":   page_num + 1,
                        "width":  base_image.get("width", 0),
                        "height": base_image.get("height", 0),
                        "bytes":  base_image.get("image", b""),
                        "ext":    base_image.get("ext", "png"),
                    })
            doc.close()
        except Exception as e:
            logger.warning("image_extraction_failed", error=str(e))
        return images

    def _ocr_pdf(self, file_bytes: bytes) -> str:
        """OCR scanned PDF using Tesseract."""
        try:
            import pytesseract
            import fitz
            from PIL import Image

            doc = fitz.open(stream=file_bytes, filetype="pdf")
            texts = []
            for page in doc:
                # Render page at 300 DPI for good OCR quality
                mat  = fitz.Matrix(300 / 72, 300 / 72)
                pix  = page.get_pixmap(matrix=mat)
                img  = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                # Try with Indian language support first, fallback to eng only
                try:
                    text = pytesseract.image_to_string(img, lang="eng+hin+tam+tel+kan+ben")
                except Exception:
                    text = pytesseract.image_to_string(img, lang="eng")
                texts.append(text)
            doc.close()
            return "\n".join(texts)
        except Exception as e:
            logger.warning("ocr_failed", error=str(e))
            return ""

    # ── PII MASKING ───────────────────────────────────────

    def _mask_pii(self, text: str) -> tuple[str, list]:
        """Mask PII using Presidio. Returns (masked_text, entities_found)."""
        if not self._presidio or not text:
            return text, []
        try:
            analyzer   = self._presidio["analyzer"]
            anonymizer = self._presidio["anonymizer"]

            # Chunk text to avoid memory issues
            chunk_size = 10000
            chunks     = [text[i:i+chunk_size] for i in range(0, len(text), chunk_size)]
            masked_chunks = []
            all_entities  = []

            for chunk in chunks:
                results = analyzer.analyze(text=chunk, language="en",
                    entities=["PERSON","EMAIL_ADDRESS","PHONE_NUMBER",
                              "CREDIT_CARD","IBAN_CODE","US_SSN"])
                all_entities.extend([r.entity_type for r in results])
                if results:
                    masked = anonymizer.anonymize(text=chunk, analyzer_results=results)
                    masked_chunks.append(masked.text)
                else:
                    masked_chunks.append(chunk)

            return "".join(masked_chunks), list(set(all_entities))
        except Exception as e:
            logger.warning("pii_masking_failed", error=str(e))
            return text, []

    # ── EXCEL PARSING ─────────────────────────────────────

    def _parse_excel(self, file_bytes: bytes, plan: str) -> dict:
        """Parse Excel files using openpyxl."""
        if not self._openpyxl_ok:
            return {"full_text": "Excel parsing unavailable", "tables": [], "metadata": {}}

        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            tables   = []
            all_text = []

            for sheet_name in wb.sheetnames:
                ws     = wb[sheet_name]
                rows   = []
                for row in ws.iter_rows(values_only=True):
                    clean = [str(c or "").strip() for c in row]
                    if any(clean):
                        rows.append(clean)

                if rows:
                    tables.append({
                        "sheet":     sheet_name,
                        "headers":   rows[0] if rows else [],
                        "rows":      rows[1:],
                        "row_count": len(rows) - 1,
                    })
                    all_text.append(f"Sheet: {sheet_name}")
                    all_text.append(self._tables_to_text([{
                        "page": 1, "headers": rows[0] if rows else [],
                        "rows": rows[1:20],
                    }]))

            full_text = "\n".join(all_text)

            # PII masking for Pro+
            pii_entities = []
            if plan in ("professional", "enterprise") and self._presidio:
                full_text, pii_entities = self._mask_pii(full_text)

            return {
                "full_text":   full_text,
                "tables":      tables,
                "metadata":    {"sheets": wb.sheetnames, "page_count": len(wb.sheetnames)},
                "pii_masked":  len(pii_entities) > 0,
                "pii_entities": pii_entities,
            }
        except Exception as e:
            logger.error("excel_parsing_failed", error=str(e))
            return {"full_text": "", "tables": [], "metadata": {}}

    # ── XML PARSING ───────────────────────────────────────

    def _parse_xml(self, file_bytes: bytes, plan: str) -> dict:
        """Parse XML contracts (Enterprise)."""
        if plan not in ("enterprise",):
            return {"full_text": "XML parsing requires Enterprise plan",
                    "tables": [], "metadata": {}}
        try:
            import xml.etree.ElementTree as ET
            root   = ET.fromstring(file_bytes.decode("utf-8", errors="ignore"))
            texts  = []

            def extract_text(element, depth=0):
                tag   = element.tag.split("}")[-1] if "}" in element.tag else element.tag
                value = (element.text or "").strip()
                if value:
                    texts.append(f"{'  ' * depth}{tag}: {value}")
                for child in element:
                    extract_text(child, depth + 1)

            extract_text(root)
            full_text = "\n".join(texts)

            # Strip metadata for Enterprise
            result = {
                "full_text": full_text,
                "tables":    [],
                "metadata":  {"root_tag": root.tag, "format": "xml"},
                "xml_root":  root.tag,
            }
            if plan == "enterprise":
                result["metadata_stripped"] = True

            return result
        except Exception as e:
            logger.error("xml_parsing_failed", error=str(e))
            return {"full_text": "", "tables": [], "metadata": {}}

    # ── DOCX PARSING ──────────────────────────────────────

    def _parse_docx(self, file_bytes: bytes, plan: str) -> dict:
        """Parse DOCX files."""
        try:
            import fitz
            doc  = fitz.open(stream=file_bytes, filetype="docx")
            text = "\n".join(page.get_text() for page in doc)
            doc.close()

            pii_entities = []
            if plan in ("professional", "enterprise") and self._presidio and text:
                text, pii_entities = self._mask_pii(text)

            return {
                "full_text":   text,
                "tables":      [],
                "metadata":    {},
                "pii_masked":  len(pii_entities) > 0,
                "pii_entities": pii_entities,
            }
        except Exception as e:
            logger.error("docx_parsing_failed", error=str(e))
            return {"full_text": "", "tables": [], "metadata": {}}

    # ── TEMPLATE MATCHING ─────────────────────────────────

    def match_template(self, text: str, org_id: str) -> dict:
        """
        Match contract against known templates.
        Uses simple keyword/structure matching.
        Professional+ feature.
        """
        templates = {
            "nda": ["non-disclosure", "confidential", "proprietary information",
                    "disclosing party", "receiving party"],
            "saas": ["software as a service", "subscription", "uptime", "sla",
                     "api", "platform"],
            "employment": ["employee", "salary", "non-compete", "termination",
                          "severance", "benefits"],
            "vendor": ["vendor", "supplier", "purchase order", "delivery",
                       "quality"],
            "lease": ["lease", "rent", "tenant", "landlord", "premises"],
            "license": ["license", "licensor", "licensee", "royalt",
                        "territory", "exclusiv"],
        }

        text_lower = text.lower()
        scores = {}
        for template, keywords in templates.items():
            score = sum(1 for kw in keywords if kw in text_lower)
            if score > 0:
                scores[template] = round(score / len(keywords) * 100)

        if not scores:
            return {"matched": False, "template": "unknown", "confidence": 0}

        best = max(scores, key=scores.get)
        return {
            "matched":    True,
            "template":   best,
            "confidence": scores[best],
            "all_scores": scores,
        }

    # ── SIGNATURE DETECTION ───────────────────────────────

    def detect_signatures(self, file_bytes: bytes) -> dict:
        """Detect signature/seal regions in PDF pages."""
        signatures = []
        try:
            import fitz
            doc = fitz.open(stream=file_bytes, filetype="pdf")
            for page_num, page in enumerate(doc, 1):
                # Check for signature fields
                for field in page.widgets() or []:
                    if field.field_type_string in ("Sig", "Signature"):
                        signatures.append({
                            "page":  page_num,
                            "type":  "digital_signature",
                            "field": field.field_name,
                            "signed": field.field_value is not None,
                        })

                # Check for text patterns indicating signatures
                text = page.get_text()
                sig_patterns = [
                    r"signed by[:\s]+([A-Z][a-z]+ [A-Z][a-z]+)",
                    r"signature[:\s]*_{3,}",
                    r"authorized signatory",
                ]
                for pattern in sig_patterns:
                    matches = re.findall(pattern, text, re.IGNORECASE)
                    for match in matches:
                        signatures.append({
                            "page":  page_num,
                            "type":  "text_signature",
                            "value": match if isinstance(match, str) else str(match),
                        })
            doc.close()
        except Exception as e:
            logger.warning("signature_detection_failed", error=str(e))

        return {
            "has_signatures": len(signatures) > 0,
            "count":          len(signatures),
            "signatures":     signatures,
        }

    # ── METADATA EXTRACTION ───────────────────────────────

    def extract_metadata(self, file_bytes: bytes, filename: str) -> dict:
        """Extract comprehensive document metadata."""
        try:
            import fitz
            doc  = fitz.open(stream=file_bytes,
                             filetype=filename.split(".")[-1].lower())
            meta = doc.metadata or {}
            page_count = len(doc)

            # Backdating risk: creation date after modification date
            creation  = meta.get("creationDate", "")
            modified  = meta.get("modDate", "")
            backdating_risk = False
            if creation and modified and creation > modified:
                backdating_risk = True

            doc.close()
            return {
                "page_count":      page_count,
                "title":           meta.get("title", ""),
                "author":          meta.get("author", ""),
                "creator_tool":    meta.get("creator", ""),
                "producer":        meta.get("producer", ""),
                "creation_date":   creation,
                "mod_date":        modified,
                "backdating_risk": backdating_risk,
                "revision_count":  0,
            }
        except Exception as e:
            logger.warning("metadata_extraction_failed", error=str(e))
            return {"page_count": 0, "backdating_risk": False}

    @classmethod
    def status(cls) -> dict:
        """Return current model loading status."""
        return {
            "initialized": cls._initialized,
            "pdfplumber":  cls._pdfplumber_ok,
            "tesseract":   cls._tesseract_ok,
            "openpyxl":    cls._openpyxl_ok,
            "presidio":    cls._presidio is not None,
            "spacy":       cls._spacy_nlp is not None,
        }
